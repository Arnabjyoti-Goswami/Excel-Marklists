import os

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side

nb_dir = os.path.abspath("")


def sort_df_column(df, sort_column):
    """
    Sorts the given dataframe w.r.t. a single column.
    The rows with the non-numeric data-types will be placed at the bottom. And the rows with the numeric data-types will be sorted in descending order.
    """
    if sort_column in df.columns:
        # Create a mask for non-numeric values
        non_numeric_mask = ~df[sort_column].apply(lambda x: isinstance(x, (int, float)))

        # Sort the DataFrame using the mask
        df_numeric = df[~non_numeric_mask].sort_values(by=sort_column, ascending=False)
        df_non_numeric = df[non_numeric_mask].sort_values(
            by=sort_column, ascending=False
        )

        # Concatenate the sorted DataFrames to combine numeric and non-numeric values
        sorted_df = pd.concat([df_numeric, df_non_numeric])

        # Reset the index to maintain a continuous index
        sorted_df.reset_index(drop=True, inplace=True)

        return sorted_df
    else:
        print(
            f"Warning: Column '{sort_column}' not found in the DataFrame. Skipping sorting..."
        )
        return df


def find_avg(df, skip_non_numeric=True, start_num_col=4):
    """
    Add a row at the bottom of the dataframe that says the average for all the columns.
    The column from where we want to start that average value columns is set to column number start_num_col. For example, if column-1 of the sheet has name, column-2 has ID, column-3 has Quiz-1 marks, and column-4 onwards are all marks of exams, then start_num_col will be 3.
    """
    start_num_col -= 1  # convert to indexing starts from 0 format for the column number starting from which we want to find the averages for
    cols = df.iloc[:, start_num_col:]
    # cols = df.iloc[:, start_num_col:].select_dtypes(include='number')  # Would not work for columns that have some few string values like 'Absent', etc., for the students that didn't give the test, so the whole column is not consistently of numeric data type.

    if skip_non_numeric:
        cols = cols.apply(
            pd.to_numeric, errors="coerce"
        )  # Convert non-numeric values to NaN
    else:
        cols = cols.apply(pd.to_numeric, errors="coerce")
        cols = cols.fillna(
            0
        )  # Replace NaN values (resulting from non-numeric values) with 0

    # Calculate the average for numeric columns (starting from 3rd column)
    avg_row = cols.mean()
    label = pd.Series(["AVERAGE"])
    avg_row = pd.concat([label, avg_row], ignore_index=True)

    # Make that series into an appropriate df to be appended to the original df
    start_col = start_num_col - 1
    avg_row = pd.DataFrame([avg_row.values], columns=df.columns[start_col:])

    # Concatenate the 2 df's
    df = pd.concat([df, avg_row], ignore_index=True)

    # Format the average values so that they are rounded to two decimal places
    df.iloc[-1, start_num_col:] = df.iloc[-1, start_num_col:].apply(
        lambda x: round(x, 2)
    )

    return df


def excel_list_to_df_dict(excel_list):
    """
    Take in a list of excel filenames (without the file extension)
    Return the a dictionary of dataframes where the keys are the filenames
    """
    dfs_dict = {}
    for file in excel_list:
        file_path = os.path.join(nb_dir, file + ".xlsx")
        df = pd.read_excel(file_path)
        dfs_dict[file] = df

    return dfs_dict


def merge_df_dict(dfs_dict):
    """
    Merge the dictionary of dataframes obtained from the excel_list_to_df_dict function into a single dataframe
    """
    concatenated_df = pd.concat(dfs_dict.values(), ignore_index=True)
    return concatenated_df


def generate_merged_excel(
    filename: str,
    excel_list: list[str],
    output_excel_sheetname: str,
    start_num_col: int,
    sort_column: None | str = None,
    skip_non_numeric: bool = False,
) -> None:
    filepath = os.path.join(nb_dir, filename + ".xlsx")

    dfs_dict = excel_list_to_df_dict(excel_list)
    df = merge_df_dict(dfs_dict)

    # Sorting must be done before adding the average values at the end
    # (else it will consider the average values as part of the values while sorting)
    if sort_column is not None:
        df = sort_df_column(df, sort_column)

    df = find_avg(df, skip_non_numeric, start_num_col)

    # Create an ExcelWriter object with 'openpyxl' engine
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            sheet_name=output_excel_sheetname,
            index=False,
            startrow=0,
            startcol=0,
        )

        workbook = writer.book
        worksheet = writer.sheets[output_excel_sheetname]

        # Adjust the column width for all columns
        for column in worksheet.columns:
            max_length = 0
            column_name = column[
                0
            ].column_letter  # Get the column name (e.g., 'A', 'B', 'C', ...)

            for cell in column:  # Find the length of the longest content in each column
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)

            adjusted_width = (
                max_length + 2
            )  # adjusts the width of all columns in the Excel worksheet to accommodate the longest content in each column. You can adjust the adjusted_width calculation to fit your specific formatting needs.
            worksheet.column_dimensions[column_name].width = adjusted_width

        # Adjust the row height and align all cells to center for all rows from 2nd to last
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True, horizontal="center", vertical="center"
                )
                worksheet.row_dimensions[cell.row].height = 20

        # Style the 1st row with a bigger height and bold text
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)  # Set text to bold
                cell.alignment = Alignment(
                    wrap_text=True, horizontal="center", vertical="center"
                )  # Center it
                worksheet.row_dimensions[
                    cell.row
                ].height = 30  # Set a bigger row height

        # Style the last row (average) with a bigger height and bold text
        for row in worksheet.iter_rows(
            min_row=worksheet.max_row, max_row=worksheet.max_row
        ):
            for cell in row:
                cell.font = Font(bold=True)  # Set text to bold
                worksheet.row_dimensions[
                    cell.row
                ].height = 30  # Set a bigger row height
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Create and apply a default black line border to all cells
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border

        # Save the final styled excel sheet
        df.to_excel(
            writer,
            sheet_name=output_excel_sheetname,
            index=False,
            startrow=0,
            startcol=0,
        )

        print(f"{filename}.xlsx saved to '{filepath}' successfully")


if __name__ == "__main__":
    long_string = input(
        "Enter the name(s) of the input excel file(s) [without .xlsx extensions, also separate the names by commas if there are multiple files]: \n"
    )
    excel_input_file_list = [s.strip() for s in long_string.split(",")]

    output_excel_filename = input(
        "Enter the name of the output excel file that you want: "
    ).strip()

    output_excel_sheetname = input(
        "Enter the name of the sheet in the output excel file that you want: "
    ).strip()

    column_to_sort_by = input(
        "Enter the name of the column in the input excel file that you want to sort the output excel file by (copy and paste the whole column heading cell of the column you want to Sort By): "
    ).strip()

    num_cols_start = input(
        "Enter the column number where the numeric values start (usually the column just after the name or the ID in marklists): "
    ).strip()

    try:
        num_cols_start = int(num_cols_start)
    except Exception as e:
        print(
            f"An error occurred while trying to convert the given input into an integer: {e}"
        )
        quit()

    skip_non_numeric_values = input(
        "Enter whether you want to skip non numeric values (like 'Absent', missing values, etc.) in the numeric columns for the calculation of average values ['y' for yes, 'n' for no] :  "
    ).strip()
    if skip_non_numeric_values not in ["y", "n", "Y", "N"]:
        print("Error! Given input is not of the correct format.")
        quit()
    skip_non_numeric_values = skip_non_numeric_values.lower()
    if skip_non_numeric_values == "y":
        skip_non_numeric_values = True
    else:
        skip_non_numeric_values == False

    generate_merged_excel(
        filename=output_excel_filename,
        excel_list=excel_input_file_list,
        output_excel_sheetname=output_excel_sheetname,
        start_num_col=num_cols_start,
        skip_non_numeric=skip_non_numeric_values,
        sort_column=column_to_sort_by,
    )
